import os
import sys
import glob
import pandas as pd
import numpy as np
import warnings
warnings.filterwarnings("ignore")
from datetime import datetime, date

# ── Paths (input data only) ──────────────────────────────────────────────────
BASE        = os.path.dirname(os.path.abspath(__file__))
ROOT        = os.path.dirname(BASE)
XLSX        = os.path.join(ROOT, "Data", "Updated Data _ 031526",
                           "Data clean and normalized-ltv computed_v2.xlsx")
MACRO_NAT   = os.path.join(ROOT, "Data", "Macro Information", "National Macro")
MACRO_STATE = os.path.join(ROOT, "Data", "Macro Information",
                            "Regional Macro", "Unemployment")

CUTOFF_DATE     = "2022-01-01"
TIER_THRESHOLDS = [0.10, 0.30, 0.60]

COL = {
    "asset_number":       5,
    "origination_date":   9,
    "original_balance":  10,
    "original_term":     11,
    "maturity_date":     12,
    "interest_rate_orig":14,
    "io_term_months":    18,
    "payment_type_code": 23,
    "is_io":             30,
    "balance_begin":     39,
    "interest_rate_curr":41,
    "balance_end":       47,
    "payment_status":    54,
    "state":             61,
    "zip_code":          62,
    "property_type":     64,
    "appraisal_orig":    70,
    "appraisal_current": 74,
    "occupancy_orig":    77,
    "noi_orig":          84,
    "dscr_orig":         88,
    "occupancy_current": 104,
    "noi_current":       109,
    "dscr_current":      112,
    "asset_left":        149,
    "cik":               150,
    "report_period":     151,
    "default_flag":      152,
}


# ═══════════════════════════════════════════════════════════════════════════════
#  STEP 21 — Loan UID Construction
# ═══════════════════════════════════════════════════════════════════════════════

def s21_excel_date_to_str(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        try:
            base = datetime(1899, 12, 30)
            return (base + pd.Timedelta(days=int(val))).strftime("%Y-%m-%d")
        except Exception:
            return str(val)
    return str(val)


def s21_load_xlsx(xlsx_path):
    import openpyxl
    print(f"Opening: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    print(f"Sheet: {ws.title}  |  max_row ~= {ws.max_row:,}")
    col_names = list(COL.keys())
    col_idxs  = list(COL.values())
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        extracted = [row[j] if j < len(row) else None for j in col_idxs]
        rows.append(extracted)
        if i % 100_000 == 0:
            print(f"  ... {i:,} rows read", flush=True)
    wb.close()
    df = pd.DataFrame(rows, columns=col_names)
    print(f"Loaded {len(df):,} rows x {len(df.columns)} columns")
    return df


def s21_clean_df(df):
    df["origination_date"] = df["origination_date"].apply(s21_excel_date_to_str)
    df["maturity_date"]    = df["maturity_date"].apply(s21_excel_date_to_str)
    num_cols = [
        "original_balance", "original_term", "interest_rate_orig",
        "io_term_months", "balance_begin", "interest_rate_curr", "balance_end",
        "appraisal_orig", "appraisal_current",
        "occupancy_orig", "noi_orig", "dscr_orig",
        "occupancy_current", "noi_current", "dscr_current",
    ]
    for c in num_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    for col in ["payment_status", "property_type", "state", "asset_left",
                "cik", "asset_number", "payment_type_code", "is_io"]:
        if col in df.columns:
            df[col] = df[col].astype(str)
    df["default_flag"] = pd.to_numeric(df["default_flag"], errors="coerce").fillna(0).astype(int)
    df["is_io"] = df["is_io"].map(lambda x: True if x is True or str(x).lower() == "true" else False)
    df["report_period"] = df["report_period"].astype(str)
    df["year"]  = df["report_period"].str[:4].apply(lambda x: int(x) if x.isdigit() else None)
    df["month"] = df["report_period"].str[5:7].apply(lambda x: int(x) if x.isdigit() else None)

    def clean_zip(x):
        if x is None or x == "" or (isinstance(x, float) and pd.isna(x)):
            return "00000"
        s = str(x).strip()
        if "-" in s:
            s = s.split("-")[0]
        try:
            return str(int(float(s))).zfill(5)
        except (ValueError, OverflowError):
            return "00000"
    df["zip_code"] = df["zip_code"].apply(clean_zip)
    df["ltv_orig"] = np.where(
        (df["appraisal_orig"] > 0) & df["appraisal_orig"].notna(),
        df["original_balance"] / df["appraisal_orig"], np.nan
    )
    df["ltv_current"] = np.where(
        (df["appraisal_current"] > 0) & df["appraisal_current"].notna(),
        df["balance_end"] / df["appraisal_current"], np.nan
    )
    return df


def s21_build_loan_uid(df):
    orig  = df["origination_date"].fillna("NA")
    mat   = df["maturity_date"].fillna("NA")
    zp    = df["zip_code"].fillna("00000")
    pt    = df["property_type"].fillna("NA").astype(str)
    bal   = pd.to_numeric(df["original_balance"], errors="coerce").fillna(0)
    bal_r = bal.round(-3).astype(int).astype(str)
    df["loan_uid"] = orig + "_" + mat + "_" + zp + "_" + pt + "_" + bal_r
    return df


def s21_classify_payment_type(row):
    is_io     = row["is_io"]
    io_months = row["io_term_months"] if pd.notna(row["io_term_months"]) else 0
    orig_term = row["original_term"]  if pd.notna(row["original_term"])  else 0
    if is_io or (io_months > 0 and orig_term > 0 and io_months >= orig_term * 0.9):
        return "Interest-Only"
    elif io_months > 0:
        return "Partial IO"
    else:
        return "Amortizing"


def run_step21():
    print("\n" + "=" * 60)
    print("  STEP 21 — LOAN UID CONSTRUCTION")
    print("=" * 60)
    df = s21_load_xlsx(XLSX)
    df = s21_clean_df(df)
    df = s21_build_loan_uid(df)
    df["payment_type"] = df.apply(s21_classify_payment_type, axis=1)

    n_uid       = df["loan_uid"].nunique()
    n_def_loans = (df.groupby("loan_uid")["default_flag"].max() == 1).sum()
    periods     = df["report_period"].dropna().unique()
    p_min = sorted(periods)[0] if len(periods) else "?"
    p_max = sorted(periods)[-1] if len(periods) else "?"
    print(f"  Unique loan_uid:           {n_uid:,}")
    print(f"  Observation period:        {p_min} to {p_max}")
    print(f"  Loans ever defaulted:      {n_def_loans:,}  ({n_def_loans/n_uid*100:.2f}%)")
    print("step21 complete.\n")
    return df


# ═══════════════════════════════════════════════════════════════════════════════
#  STEP 22 — EDA
# ═══════════════════════════════════════════════════════════════════════════════

def s22_classify_payment_type(row):
    is_io     = row.get("is_io", False)
    io_months = row.get("io_term_months", 0) or 0
    orig_term = row.get("original_term", 0) or 0
    try:
        io_months = float(io_months)
        orig_term = float(orig_term)
    except Exception:
        return "Other"
    if is_io or (io_months > 0 and orig_term > 0 and io_months >= orig_term * 0.9):
        return "Interest-Only"
    elif io_months > 0:
        return "Partial IO"
    elif orig_term > 0 or not is_io:
        return "Amortizing"
    return "Other"


def s22_build_loan_df(df):
    first = df.sort_values("report_period").groupby("loan_uid").first().reset_index()
    ever_default = df.groupby("loan_uid")["default_flag"].max().rename("ever_defaulted")
    def_rows = df[df["default_flag"] == 1].copy()
    first_default_period = def_rows.groupby("loan_uid")["report_period"].min().rename("first_default_period")
    first_default_year   = def_rows.groupby("loan_uid")["year"].min().rename("first_default_year")
    loan_df = first.join(ever_default, on="loan_uid")
    loan_df = loan_df.join(first_default_period, on="loan_uid")
    loan_df = loan_df.join(first_default_year, on="loan_uid")
    loan_df["ever_defaulted"] = loan_df["ever_defaulted"].fillna(0).astype(int)
    return loan_df


def s22_eda1_overview(df, loan_df):
    n_obs      = len(df)
    n_loans    = loan_df["loan_uid"].nunique()
    p_min      = df["report_period"].dropna().min()
    p_max      = df["report_period"].dropna().max()
    n_def      = loan_df["ever_defaulted"].sum()
    def_rate   = n_def / n_loans * 100 if n_loans > 0 else 0
    lines = [
        "\n" + "=" * 58,
        "  EDA-1: Basic Dataset Overview",
        "=" * 58,
        f"  Total loan-month observations:     {n_obs:>12,}",
        f"  Total unique loans (loan_uid):      {n_loans:>12,}",
        f"  Observation period:          {p_min}  to  {p_max}",
        f"  Avg months observed per loan:       {n_obs/n_loans:>12.1f}",
        f"  Loans ever defaulted:               {n_def:>12,}",
        f"  Loans never defaulted:              {n_loans - n_def:>12,}",
        f"  Overall loan-level default rate:    {def_rate:>11.2f}%",
        "=" * 58,
    ]
    return "\n".join(lines)


def s22_eda2_annual_default(df, loan_df):
    target_years = list(range(2016, 2026))
    rows = []
    for yr in target_years:
        active     = set(df[df["year"] == yr]["loan_uid"].unique())
        first_def  = loan_df[loan_df["first_default_year"] == yr]["loan_uid"]
        n_active   = len(active)
        n_firstdef = len(first_def)
        rate       = n_firstdef / n_active * 100 if n_active > 0 else 0
        rows.append({"Year": yr, "Active Loans": n_active,
                     "First-Time Defaults": n_firstdef, "Default Rate (%)": round(rate, 2)})
    result = pd.DataFrame(rows)
    col_w  = [6, 14, 21, 16]
    sep    = "-" * (sum(col_w) + 9)
    hdr    = f"{'Year':<{col_w[0]}} | {'Active Loans':>{col_w[1]}} | {'First-Time Defaults':>{col_w[2]}} | {'Default Rate (%)':>{col_w[3]}}"
    lines  = ["\n" + "=" * 65, "  EDA-2: Annual Default Rate (2016–2025, First-Time Defaults)", "=" * 65, hdr, sep]
    for _, r in result.iterrows():
        lines.append(
            f"{int(r['Year']):<{col_w[0]}} | {int(r['Active Loans']):>{col_w[1]},} | "
            f"{int(r['First-Time Defaults']):>{col_w[2]},} | {r['Default Rate (%)']:>{col_w[3]}.2f}%"
        )
    lines.append("=" * 65)
    peak = result.loc[result["Default Rate (%)"].idxmax()]
    lines += ["", "  INTERPRETATION:",
              f"  - Peak first-time default year: {int(peak['Year'])} ({peak['Default Rate (%)']:.1f}%)",
              "  - 2020: COVID-19 shock caused sharp spike in lodging/retail defaults.",
              "  - 2022-2024: Rising refinancing costs created stress for IO loans."]
    return "\n".join(lines)


def s22_eda3_payment_type(loan_df):
    pt_order = ["Interest-Only", "Partial IO", "Amortizing", "Other"]
    loan_df  = loan_df.copy()
    if "payment_type" not in loan_df.columns:
        loan_df["payment_type"] = loan_df.apply(s22_classify_payment_type, axis=1)
    total = len(loan_df)
    rows  = []
    for pt in pt_order:
        sub   = loan_df[loan_df["payment_type"] == pt]
        n     = len(sub)
        n_def = sub["ever_defaulted"].sum()
        rows.append({"Payment Type": pt, "Unique Loans": n,
                     "% of Total": n/total*100 if total>0 else 0,
                     "Defaulted": n_def, "Default Rate (%)": n_def/n*100 if n>0 else 0})
    col_w = [16, 13, 10, 10, 16]
    hdr   = (f"{'Payment Type':<{col_w[0]}} | {'Unique Loans':>{col_w[1]}} | "
             f"{'% of Total':>{col_w[2]}} | {'Defaulted':>{col_w[3]}} | {'Default Rate (%)':>{col_w[4]}}")
    sep   = "-" * (sum(col_w) + 12)
    lines = ["\n" + "=" * 75, "  EDA-3: Payment Type Breakdown (Loan-Level)", "=" * 75, hdr, sep]
    for r in rows:
        lines.append(
            f"{r['Payment Type']:<{col_w[0]}} | {r['Unique Loans']:>{col_w[1]},} | "
            f"{r['% of Total']:>{col_w[2]}.1f}% | {r['Defaulted']:>{col_w[3]},} | "
            f"{r['Default Rate (%)']:>{col_w[4]}.2f}%"
        )
    lines.append("=" * 75)
    return "\n".join(lines)


def s22_eda4_annual_by_payment_type(df, loan_df):
    target_years = list(range(2016, 2026))
    pt_cats      = ["Interest-Only", "Partial IO", "Amortizing"]
    if "payment_type" not in loan_df.columns:
        loan_df = loan_df.copy()
        loan_df["payment_type"] = loan_df.apply(s22_classify_payment_type, axis=1)
    uid_to_pt  = loan_df.set_index("loan_uid")["payment_type"].to_dict()
    uid_to_fdy = loan_df.set_index("loan_uid")["first_default_year"].to_dict()
    rows = []
    for yr in target_years:
        active_uids = set(df[df["year"] == yr]["loan_uid"].unique())
        row = {"Year": yr}
        for pt in pt_cats:
            active_pt = {u for u in active_uids if uid_to_pt.get(u) == pt}
            def_pt    = {u for u in active_pt if uid_to_fdy.get(u) == yr}
            row[pt]   = (len(def_pt) / len(active_pt) * 100) if active_pt else 0
        rows.append(row)
    result = pd.DataFrame(rows)
    col_w  = [6, 15, 12, 12]
    hdr    = f"{'Year':<{col_w[0]}} | {'Interest-Only':>{col_w[1]}} | {'Partial IO':>{col_w[2]}} | {'Amortizing':>{col_w[3]}}"
    sep    = "-" * (sum(col_w) + 9)
    lines  = ["\n" + "=" * 55, "  EDA-4: Annual Default Rate by Payment Type (%)", "=" * 55, hdr, sep]
    for _, r in result.iterrows():
        lines.append(
            f"{int(r['Year']):<{col_w[0]}} | {r['Interest-Only']:>{col_w[1]}.1f}% | "
            f"{r['Partial IO']:>{col_w[2]}.1f}% | {r['Amortizing']:>{col_w[3]}.1f}%"
        )
    lines.append("=" * 55)
    return "\n".join(lines)


def s22_eda5_state(loan_df, min_loans=10):
    by_state = (
        loan_df.groupby("state")
        .agg(total_loans=("loan_uid", "count"), defaulted=("ever_defaulted", "sum"))
        .reset_index()
    )
    by_state["default_rate"] = by_state["defaulted"] / by_state["total_loans"] * 100
    by_state = by_state[by_state["total_loans"] >= min_loans].sort_values("default_rate", ascending=False)
    col_w = [6, 14, 13, 16]
    hdr   = (f"{'State':<{col_w[0]}} | {'Total Loans':>{col_w[1]}} | "
             f"{'Defaulted':>{col_w[2]}} | {'Default Rate (%)':>{col_w[3]}}")
    sep   = "-" * (sum(col_w) + 9)
    lines = ["\n" + "=" * 58, f"  EDA-5: State-Level Default Analysis (>={min_loans} loans)", "=" * 58, hdr, sep]
    for _, r in by_state.iterrows():
        lines.append(
            f"{str(r['state']):<{col_w[0]}} | {int(r['total_loans']):>{col_w[1]},} | "
            f"{int(r['defaulted']):>{col_w[2]},} | {r['default_rate']:>{col_w[3]}.2f}%"
        )
    lines.append("=" * 58)
    return "\n".join(lines)


def s22_eda6_property_type(loan_df):
    by_pt = (
        loan_df.groupby("property_type")
        .agg(total_loans=("loan_uid", "count"), defaulted=("ever_defaulted", "sum"))
        .reset_index()
    )
    by_pt["default_rate"] = by_pt["defaulted"] / by_pt["total_loans"] * 100
    by_pt = by_pt.sort_values("default_rate", ascending=False)
    col_w = [15, 13, 11, 16]
    hdr   = (f"{'Property Type':<{col_w[0]}} | {'Unique Loans':>{col_w[1]}} | "
             f"{'Defaulted':>{col_w[2]}} | {'Default Rate (%)':>{col_w[3]}}")
    sep   = "-" * (sum(col_w) + 9)
    lines = ["\n" + "=" * 65, "  EDA-6: Property Type Default Analysis (Loan-Level)", "=" * 65, hdr, sep]
    for _, r in by_pt.iterrows():
        lines.append(
            f"{str(r['property_type']):<{col_w[0]}} | {int(r['total_loans']):>{col_w[1]},} | "
            f"{int(r['defaulted']):>{col_w[2]},} | {r['default_rate']:>{col_w[3]}.2f}%"
        )
    lines.append("=" * 65)
    return "\n".join(lines)


def run_step22(panel_df):
    print("\n" + "=" * 58)
    print("  STEP 22 — EDA")
    print("=" * 58)
    df      = panel_df
    loan_df = s22_build_loan_df(df)
    if "payment_type" not in loan_df.columns:
        loan_df["payment_type"] = loan_df.apply(s22_classify_payment_type, axis=1)
    if "payment_type" not in df.columns:
        uid_to_pt = loan_df.set_index("loan_uid")["payment_type"]
        df = df.copy()
        df["payment_type"] = df["loan_uid"].map(uid_to_pt)

    print(s22_eda1_overview(df, loan_df))
    print(s22_eda2_annual_default(df, loan_df))
    print(s22_eda3_payment_type(loan_df))
    print(s22_eda4_annual_by_payment_type(df, loan_df))
    print(s22_eda5_state(loan_df))
    print(s22_eda6_property_type(loan_df))
    print("step22 complete.\n")


# ═══════════════════════════════════════════════════════════════════════════════
#  STEP 24 — Feature Engineering
# ═══════════════════════════════════════════════════════════════════════════════

def s24_load_national_macro():
    import openpyxl
    series = {}

    dgs10_path = os.path.join(MACRO_NAT, "DGS10.xlsx")
    if os.path.exists(dgs10_path):
        wb = openpyxl.load_workbook(dgs10_path, read_only=True)
        ws = wb["Daily"]
        rows = [r for r in ws.iter_rows(values_only=True)]
        wb.close()
        daily = pd.DataFrame(rows[1:], columns=["date", "dgs10"])
        daily["date"] = pd.to_datetime(daily["date"], errors="coerce")
        daily = daily.dropna(subset=["date"])
        daily["dgs10"] = pd.to_numeric(daily["dgs10"], errors="coerce")
        monthly = daily.set_index("date")["dgs10"].resample("MS").mean().reset_index()
        monthly["ym"] = monthly["date"].dt.strftime("%Y-%m")
        series["dgs10"] = monthly[["ym", "dgs10"]].rename(columns={"dgs10": "nat_10y_rate"})

    unrate_path = os.path.join(MACRO_NAT, "UNRATE.xlsx")
    if os.path.exists(unrate_path):
        wb = openpyxl.load_workbook(unrate_path, read_only=True)
        ws = wb[wb.sheetnames[-1]]
        rows = [r for r in ws.iter_rows(values_only=True)]
        wb.close()
        m = pd.DataFrame(rows[1:], columns=["date", "unrate"])
        m["date"] = pd.to_datetime(m["date"], errors="coerce")
        m = m.dropna(subset=["date"])
        m["unrate"] = pd.to_numeric(m["unrate"], errors="coerce")
        m["ym"] = m["date"].dt.strftime("%Y-%m")
        series["unrate"] = m[["ym", "unrate"]].rename(columns={"unrate": "nat_unrate"})

    gdp_path = os.path.join(MACRO_NAT, "GDP.xlsx")
    if os.path.exists(gdp_path):
        wb = openpyxl.load_workbook(gdp_path, read_only=True)
        ws = wb[wb.sheetnames[-1]]
        rows = [r for r in ws.iter_rows(values_only=True)]
        wb.close()
        q = pd.DataFrame(rows[1:], columns=["date", "gdp"])
        q["date"] = pd.to_datetime(q["date"], errors="coerce")
        q = q.dropna(subset=["date"])
        q["gdp"] = pd.to_numeric(q["gdp"], errors="coerce")
        idx = pd.date_range(q["date"].min(), q["date"].max() + pd.DateOffset(months=3), freq="MS")
        q = q.set_index("date").reindex(idx).ffill().reset_index()
        q.columns = ["date", "gdp"]
        q["ym"] = q["date"].dt.strftime("%Y-%m")
        q["gdp_yoy"] = q["gdp"].pct_change(12) * 100
        series["gdp"] = q[["ym", "gdp_yoy"]].rename(columns={"gdp_yoy": "nat_gdp_yoy"})

    macro_nat = None
    for name, s in series.items():
        macro_nat = s if macro_nat is None else macro_nat.merge(s, on="ym", how="outer")
    return macro_nat


def s24_load_state_macro():
    import openpyxl, re
    state_abbrev = {
        "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
        "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
        "Florida": "FL", "Georgia": "GA", "Hawaii": "HI", "Idaho": "ID",
        "Illinois": "IL", "Indiana": "IN", "Iowa": "IA", "Kansas": "KS",
        "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME", "Maryland": "MD",
        "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN",
        "Mississippi": "MS", "Missouri": "MO", "Montana": "MT", "Nebraska": "NE",
        "Nevada": "NV", "New Hampshire": "NH", "New Jersey": "NJ",
        "New Mexico": "NM", "New York": "NY", "North Carolina": "NC",
        "North Dakota": "ND", "Ohio": "OH", "Oklahoma": "OK", "Oregon": "OR",
        "Pennsylvania": "PA", "Rhode Island": "RI", "South Carolina": "SC",
        "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX", "Utah": "UT",
        "Vermont": "VT", "Virginia": "VA", "Washington": "WA",
        "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
        "the District of Columbia": "DC", "District of Columbia": "DC",
    }
    files   = glob.glob(os.path.join(MACRO_STATE, "Unemployment Rate in *.xlsx"))
    records = []
    for fpath in files:
        fname  = os.path.basename(fpath)
        match  = re.search(r"Unemployment Rate in (.+?)\.xlsx", fname)
        if not match: continue
        abbrev = state_abbrev.get(match.group(1).strip())
        if not abbrev: continue
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True)
            ws = wb["Monthly"]
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception:
            continue
        for r in rows[1:]:
            if r[0] is None: continue
            date_val = pd.to_datetime(r[0], errors="coerce")
            val      = pd.to_numeric(r[1], errors="coerce") if len(r) > 1 else np.nan
            if pd.isna(date_val): continue
            records.append({"ym": date_val.strftime("%Y-%m"), "state": abbrev, "state_unrate": val})
    if not records:
        return pd.DataFrame(columns=["ym", "state", "state_unrate"])
    return pd.DataFrame(records)


def s24_aggregate_to_loan_period(df):
    df = df.copy()
    df["ac_key"] = df["asset_left"].astype(str) + "_" + df["cik"].astype(str)

    def modal(s):
        m = s.mode()
        return m.iloc[0] if len(m) > 0 else np.nan

    agg_dict = {
        "loan_uid": "first", "asset_left": "first", "cik": "first",
        "year": "first", "month": "first",
        "default_flag": "max", "payment_status": "first",
        "origination_date": "first", "maturity_date": "first",
        "original_term": "first", "io_term_months": "first",
        "is_io": "first", "payment_type": "first",
        "interest_rate_orig": "mean", "interest_rate_curr": "mean",
        "original_balance": "sum", "balance_begin": "sum", "balance_end": "sum",
        "appraisal_orig": "sum", "appraisal_current": "sum",
        "noi_orig": "sum", "noi_current": "sum",
        "occupancy_orig": "mean", "occupancy_current": "mean",
        "dscr_orig": "mean", "dscr_current": "mean",
        "state": "first", "zip_code": "first", "property_type": modal,
    }
    agg_dict = {k: v for k, v in agg_dict.items() if k in df.columns}
    grouped  = df.groupby(["ac_key", "report_period"], sort=False).agg(agg_dict).reset_index()
    print(f"  Aggregated {len(df):,} property rows -> {len(grouped):,} loan-period rows")
    print(f"  Unique loans (ac_key): {grouped['ac_key'].nunique():,}")
    return grouped


def s24_engineer_features(df, macro_nat, state_macro):
    df = df.sort_values(["ac_key", "report_period"]).copy()

    df["log_orig_balance"]    = np.log1p(pd.to_numeric(df["original_balance"], errors="coerce").clip(lower=0))
    df["log_current_balance"] = np.log1p(pd.to_numeric(df["balance_end"], errors="coerce").clip(lower=0))
    orig_bal = pd.to_numeric(df["original_balance"], errors="coerce")
    end_bal  = pd.to_numeric(df["balance_end"], errors="coerce")
    df["balance_paiddown"] = np.where(orig_bal > 0, (orig_bal - end_bal) / orig_bal, 0)

    df["ym_dt"]   = pd.to_datetime(df["report_period"] + "-01", errors="coerce")
    df["orig_dt"] = pd.to_datetime(df["origination_date"], errors="coerce")
    df["mat_dt"]  = pd.to_datetime(df["maturity_date"], errors="coerce")
    df["loan_age_months"]    = ((df["ym_dt"] - df["orig_dt"]) / pd.Timedelta("30.44D")).clip(lower=0)
    df["months_to_maturity"] = ((df["mat_dt"] - df["ym_dt"]) / pd.Timedelta("30.44D")).clip(lower=0)

    pt_map = {"Interest-Only": 3, "Partial IO": 2, "Amortizing": 1, "Other": 0}
    df["payment_type_ord"] = df["payment_type"].map(pt_map).fillna(0)

    ir_curr = pd.to_numeric(df["interest_rate_curr"], errors="coerce")
    ir_orig = pd.to_numeric(df["interest_rate_orig"], errors="coerce")
    df["interest_rate"] = ir_curr.fillna(ir_orig)

    appr_curr = pd.to_numeric(df["appraisal_current"], errors="coerce")
    appr_orig = pd.to_numeric(df["appraisal_orig"], errors="coerce")
    ltv_curr      = np.where((appr_curr > 0) & appr_curr.notna(), end_bal / appr_curr, np.nan)
    ltv_orig_calc = np.where((appr_orig > 0) & appr_orig.notna(), orig_bal / appr_orig, np.nan)
    df["ltv_feature"] = np.where(~np.isnan(ltv_curr), ltv_curr, ltv_orig_calc)

    dscr_curr = pd.to_numeric(df["dscr_current"], errors="coerce")
    dscr_orig = pd.to_numeric(df["dscr_orig"], errors="coerce")
    df["dscr_feature"] = dscr_curr.fillna(dscr_orig)

    df["dscr_lag1"]   = df.groupby("ac_key")["dscr_feature"].shift(1)
    df["dscr_3m_chg"] = (
        df.groupby("ac_key")["dscr_lag1"].transform(lambda x: x.rolling(3, min_periods=1).mean())
        - df["dscr_lag1"]
    )
    df["ltv_lag1"]   = df.groupby("ac_key")["ltv_feature"].shift(1)
    df["ltv_3m_chg"] = (
        df.groupby("ac_key")["ltv_lag1"].transform(lambda x: x.rolling(3, min_periods=1).mean())
        - df["ltv_lag1"]
    )

    occ_curr = pd.to_numeric(df["occupancy_current"], errors="coerce")
    occ_orig = pd.to_numeric(df["occupancy_orig"], errors="coerce")
    df["occupancy"] = occ_curr.fillna(occ_orig)

    noi_curr = pd.to_numeric(df["noi_current"], errors="coerce")
    noi_orig = pd.to_numeric(df["noi_orig"], errors="coerce")
    df["log_noi"]       = np.log1p(noi_curr.fillna(noi_orig).clip(lower=0))
    df["log_appraisal"] = np.log1p(appr_curr.fillna(appr_orig).clip(lower=0))

    prop_dummies = pd.get_dummies(df["property_type"], prefix="pt", drop_first=False)
    df = pd.concat([df, prop_dummies], axis=1)

    df["prior_default_flag"] = df.groupby("ac_key")["default_flag"].shift(1).fillna(0)
    df["prior_default_ever"] = df.groupby("ac_key")["prior_default_flag"].cummax()
    df["def_lag1"]           = df.groupby("ac_key")["default_flag"].shift(1).fillna(0)
    df["n_def_12m"] = (
        df.groupby("ac_key")["def_lag1"].transform(lambda x: x.rolling(12, min_periods=1).sum())
    )

    def months_since_last_def(series):
        result, last = [], np.nan
        for i, val in enumerate(series):
            if val == 1:
                last = i
            result.append(i - last if not np.isnan(last) else 0)
        return result

    df["months_since_def"] = (
        df.groupby("ac_key")["def_lag1"]
        .transform(lambda x: pd.Series(months_since_last_def(x.values), index=x.index))
    )

    df["ym"] = df["report_period"]

    # Macro with 3-month lag: use macro data from 3 months before report_period
    df["ym_macro"] = (
        pd.to_datetime(df["report_period"] + "-01") - pd.DateOffset(months=3)
    ).dt.strftime("%Y-%m")

    if macro_nat is not None and len(macro_nat) > 0:
        df = df.merge(macro_nat.rename(columns={"ym": "ym_macro"}), on="ym_macro", how="left")
    else:
        df["nat_10y_rate"] = np.nan
        df["nat_unrate"]   = np.nan
        df["nat_gdp_yoy"]  = np.nan

    if state_macro is not None and len(state_macro) > 0:
        df = df.merge(
            state_macro.rename(columns={"ym": "ym_macro"}),
            on=["ym_macro", "state"], how="left"
        )
    else:
        df["state_unrate"] = np.nan

    df = df.drop(columns=["ym_macro"], errors="ignore")

    df["year_feat"]    = pd.to_numeric(df["year"], errors="coerce")
    df["month_feat"]   = pd.to_numeric(df["month"], errors="coerce")
    df["quarter_feat"] = ((df["month_feat"] - 1) // 3 + 1)

    def macro_regime(ym):
        if ym < "2022-03": return 0
        elif ym <= "2023-07": return 1
        else: return 2

    df["macro_regime"] = df["ym"].apply(macro_regime)
    df["covid_period"] = df["ym"].apply(lambda x: 1 if "2020-03" <= x <= "2021-12" else 0)

    # 6-month forward default target
    df = df.sort_values(["ac_key", "report_period"]).copy()
    fwd = pd.Series(0, index=df.index, dtype=float)
    for h in range(1, 7):
        fwd = fwd.combine(
            df.groupby("ac_key")["default_flag"].shift(-h).fillna(0), max
        )
    df["target"] = fwd.astype(int)

    # Drop last observation of each loan (no future data to label)
    df["last_month"] = df.groupby("ac_key")["report_period"].transform("max")
    df = df[df["report_period"] < df["last_month"]].drop(columns=["last_month"])

    return df


FEATURE_COLS = [
    "log_orig_balance", "log_current_balance", "balance_paiddown",
    "loan_age_months", "months_to_maturity", "payment_type_ord", "interest_rate",
    "ltv_feature", "dscr_feature", "dscr_3m_chg", "ltv_3m_chg",
    "occupancy", "log_appraisal", "log_noi",
    "prior_default_ever", "n_def_12m", "months_since_def",
    "nat_10y_rate", "nat_unrate", "nat_gdp_yoy", "state_unrate",
    "year_feat", "month_feat", "quarter_feat", "macro_regime", "covid_period",
]


def run_step24(panel_df):
    print("\n" + "=" * 60)
    print("  STEP 24 — FEATURE ENGINEERING")
    print("=" * 60)

    print("\nAggregating property rows to loan-period level...")
    df = s24_aggregate_to_loan_period(panel_df)

    print("\nLoading national macro data...")
    macro_nat = s24_load_national_macro()
    if macro_nat is not None:
        print(f"National macro: {len(macro_nat)} months")
    else:
        print("WARNING: National macro data not found")

    print("\nLoading state unemployment data...")
    state_macro = s24_load_state_macro()
    if len(state_macro) > 0:
        print(f"State unemployment: {len(state_macro):,} records, {state_macro['state'].nunique()} states")
    else:
        print("WARNING: State macro data not found")

    print("\nEngineering features...")
    df = s24_engineer_features(df, macro_nat, state_macro)

    pos_rate = df["target"].mean() * 100
    print(f"  Total loan-period observations: {len(df):,}")
    print(f"  Positive (default) rate:        {pos_rate:.2f}%")

    pt_dummies  = [c for c in df.columns if c.startswith("pt_")]
    all_feats   = FEATURE_COLS + pt_dummies
    avail_feats = [c for c in all_feats if c in df.columns]

    out_df = df[["ac_key", "loan_uid", "report_period", "year", "state",
                 "target"] + avail_feats].copy()
    print(f"\nFeature matrix: {len(out_df):,} rows x {len(avail_feats)} features")
    print("step24 complete.\n")
    return out_df


# ═══════════════════════════════════════════════════════════════════════════════
#  STEP 25 — XGBoost Modeling
# ═══════════════════════════════════════════════════════════════════════════════

def s25_train_test_split(df, cutoff=CUTOFF_DATE):
    loan_first  = df.groupby("ac_key")["report_period"].min()
    train_loans = loan_first[loan_first < cutoff].index
    test_loans  = loan_first[loan_first >= cutoff].index
    train = df[df["ac_key"].isin(train_loans)].copy()
    test  = df[df["ac_key"].isin(test_loans)].copy()
    total = len(train_loans) + len(test_loans)
    print("\n=== TRAIN/TEST SPLIT ===")
    print(f"  Cutoff date:           {cutoff}")
    print(f"  Train loans:           {len(train_loans):,}  ({len(train_loans)/total*100:.1f}%)")
    print(f"  Test loans:            {len(test_loans):,}  ({len(test_loans)/total*100:.1f}%)")
    print(f"  Train positive rate:   {train['target'].mean()*100:.2f}%")
    print(f"  Test positive rate:    {test['target'].mean()*100:.2f}%")
    train_pct = len(train_loans) / total * 100
    if train_pct < 70 or train_pct > 90:
        print(f"  NOTE: Train = {train_pct:.1f}% — consider adjusting CUTOFF_DATE")
    return train, test, train_loans, test_loans


def s25_carve_validation(train, val_fraction=0.10):
    loan_first = train.groupby("ac_key")["report_period"].min().sort_values()
    n_val      = max(1, int(len(loan_first) * val_fraction))
    val_loans  = loan_first.iloc[-n_val:].index
    tr_loans   = loan_first.iloc[:-n_val].index
    tr  = train[train["ac_key"].isin(tr_loans)].copy()
    val = train[train["ac_key"].isin(val_loans)].copy()
    print(f"\n  Validation carved: {len(val_loans):,} loans, {len(val):,} obs (positive: {val['target'].mean()*100:.2f}%)")
    return tr, val


def s25_get_feature_cols(df):
    exclude = {"ac_key", "loan_uid", "report_period", "year", "state", "target",
               "ym_dt", "orig_dt", "mat_dt", "ym", "last_month", "fwd_cutoff"}
    return [c for c in df.columns if c not in exclude]


def s25_target_encode_state(train, test, val):
    state_mean  = train.groupby("state")["target"].mean().rename("state_te")
    global_mean = train["target"].mean()
    for part in [train, test, val]:
        part["state_te"] = part["state"].map(state_mean).fillna(global_mean)
    return train, test, val


def s25_impute_features(train, test, val, feat_cols):
    medians = train[feat_cols].median()
    for part in [train, test, val]:
        for c in feat_cols:
            if c in part.columns:
                part[c] = part[c].fillna(medians[c])
    return train, test, val


def s25_train_xgboost(X_tr, y_tr, X_val, y_val):
    try:
        import xgboost as xgb
    except ImportError:
        print("ERROR: xgboost not installed. Run: pip install xgboost")
        sys.exit(1)
    scale_pos_weight = len(y_tr[y_tr == 0]) / max(len(y_tr[y_tr == 1]), 1)
    print(f"\n  scale_pos_weight = {scale_pos_weight:.1f}")
    model = xgb.XGBClassifier(
        n_estimators=500,
        max_depth=6,
        learning_rate=0.05,
        subsample=0.8,
        colsample_bytree=0.8,
        scale_pos_weight=scale_pos_weight,
        eval_metric="aucpr",
        early_stopping_rounds=50,
        random_state=42,
        n_jobs=-1,
        verbosity=0,
    )
    model.fit(X_tr, y_tr, eval_set=[(X_val, y_val)], verbose=50)
    print(f"  Best iteration: {model.best_iteration}")
    return model


def s25_evaluate_model(model, X_test, y_test, feat_cols):
    from sklearn.metrics import (
        roc_auc_score, average_precision_score,
        precision_score, recall_score, f1_score, confusion_matrix,
        precision_recall_curve
    )
    proba   = model.predict_proba(X_test)[:, 1]
    roc_auc = roc_auc_score(y_test, proba)
    pr_auc  = average_precision_score(y_test, proba)

    pred_05 = (proba >= 0.5).astype(int)
    prec_05 = precision_score(y_test, pred_05, zero_division=0)
    rec_05  = recall_score(y_test, pred_05, zero_division=0)
    f1_05   = f1_score(y_test, pred_05, zero_division=0)

    precisions, recalls, thresholds = precision_recall_curve(y_test, proba)
    f1s     = 2 * precisions * recalls / (precisions + recalls + 1e-9)
    best_i  = np.argmax(f1s)
    best_th = thresholds[best_i] if best_i < len(thresholds) else 0.5
    pred_bt = (proba >= best_th).astype(int)
    prec_bt = precision_score(y_test, pred_bt, zero_division=0)
    rec_bt  = recall_score(y_test, pred_bt, zero_division=0)
    f1_bt   = f1_score(y_test, pred_bt, zero_division=0)

    cm = confusion_matrix(y_test, pred_bt)

    lines = [
        "\n" + "=" * 55,
        "  MODEL EVALUATION (Test Set)",
        "=" * 55,
        f"  ROC-AUC:                  {roc_auc:.4f}",
        f"  PR-AUC (Avg Precision):   {pr_auc:.4f}  <- primary metric",
        f"  Precision @ 0.5:          {prec_05:.4f}",
        f"  Recall    @ 0.5:          {rec_05:.4f}",
        f"  F1        @ 0.5:          {f1_05:.4f}",
        f"  Best threshold (max F1):  {best_th:.3f}",
        f"  Precision @ best thresh:  {prec_bt:.4f}",
        f"  Recall    @ best thresh:  {rec_bt:.4f}",
        f"  F1        @ best thresh:  {f1_bt:.4f}",
        "",
        "  CONFUSION MATRIX (@ best threshold)",
        f"                   Predicted 0   Predicted 1",
        f"  Actual 0         {cm[0,0]:>11,}   {cm[0,1]:>11,}",
        f"  Actual 1         {cm[1,0]:>11,}   {cm[1,1]:>11,}",
        "=" * 55,
    ]
    print("\n".join(lines))

    try:
        scores = model.get_booster().get_score(importance_type="gain")
        imp_df = pd.DataFrame(list(scores.items()), columns=["feature", "gain"])
        imp_df = imp_df.sort_values("gain", ascending=False).head(20)
    except Exception:
        imp_df = pd.DataFrame({
            "feature": feat_cols,
            "importance": model.feature_importances_
        }).sort_values("importance", ascending=False).head(20)
        imp_df = imp_df.rename(columns={"importance": "gain"})

    print("\n  TOP 20 FEATURES BY XGBoost GAIN IMPORTANCE:")
    print("  " + "-" * 50)
    for rank, (_, r) in enumerate(imp_df.iterrows(), 1):
        print(f"  {rank:>2}. {r['feature']:<35} {r['gain']:>10.1f}")

    try:
        import shap
        explainer   = shap.TreeExplainer(model)
        shap_values = explainer.shap_values(X_test[:1000])
        mean_abs    = np.abs(shap_values).mean(axis=0)
        shap_df     = pd.DataFrame({"feature": feat_cols, "mean_abs_shap": mean_abs})
        shap_df     = shap_df.sort_values("mean_abs_shap", ascending=False).head(10)
        print("\n  TOP 10 FEATURES BY MEAN |SHAP| (sample of 1000):")
        print("  " + "-" * 50)
        for _, r in shap_df.iterrows():
            print(f"  {r['feature']:<35} {r['mean_abs_shap']:>10.4f}")
    except ImportError:
        print("\n  (SHAP not installed — skipping)")

    return proba, roc_auc, pr_auc, best_th


def s25_print_risk_tiers(proba, y_test):
    def assign_tier(p):
        if p < TIER_THRESHOLDS[0]:   return "Tier 1 — Low Risk"
        elif p < TIER_THRESHOLDS[1]: return "Tier 2 — Watch"
        elif p < TIER_THRESHOLDS[2]: return "Tier 3 — High Risk"
        else:                        return "Tier 4 — Critical"

    tiers   = [assign_tier(p) for p in proba]
    tier_df = pd.DataFrame({"tier": tiers, "actual": y_test.values})
    tier_order = ["Tier 1 — Low Risk", "Tier 2 — Watch",
                  "Tier 3 — High Risk", "Tier 4 — Critical"]
    total = len(tier_df)
    lines = [
        "\n" + "=" * 78,
        "  RISK TIER DISTRIBUTION (Test Set)",
        "=" * 78,
        f"  {'Tier':<22} | {'Count':>8} | {'% of Total':>10} | {'Actual Default Rate':>20}",
        "  " + "-" * 68,
    ]
    for t in tier_order:
        sub = tier_df[tier_df["tier"] == t]
        n   = len(sub)
        pct = n / total * 100 if total > 0 else 0
        dr  = sub["actual"].mean() * 100 if n > 0 else 0
        lines.append(f"  {t:<22} | {n:>8,} | {pct:>9.1f}% | {dr:>19.1f}%")
    lines += [
        "=" * 78,
        "",
        "  RECOMMENDED LENDER ACTIONS BY TIER:",
        "  Tier 1 — Low Risk  : Routine monitoring, standard reporting cycle.",
        "  Tier 2 — Watch     : Enhanced monitoring, quarterly credit review.",
        "  Tier 3 — High Risk : Active workout planning, reserve increase,",
        "                       monthly reporting required.",
        "  Tier 4 — Critical  : Immediate intervention, escalate to special",
        "                       servicing, initiate borrower contact.",
    ]
    print("\n".join(lines))


def run_step25(features_df):
    print("\n" + "=" * 55)
    print("  STEP 25 — XGBOOST MODELING PIPELINE")
    print(f"  Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 55)

    df        = features_df
    feat_cols = s25_get_feature_cols(df)
    train, test, _, _ = s25_train_test_split(df)
    train_tr, train_val = s25_carve_validation(train)
    train_tr, test, train_val = s25_target_encode_state(train_tr, test, train_val)
    feat_cols = [c for c in feat_cols if c != "state"]
    if "state_te" not in feat_cols and "state_te" in train_tr.columns:
        feat_cols.append("state_te")
    train_tr, test, train_val = s25_impute_features(train_tr, test, train_val, feat_cols)

    X_tr  = train_tr[feat_cols]
    y_tr  = train_tr["target"]
    X_val = train_val[feat_cols]
    y_val = train_val["target"]
    X_te  = test[feat_cols]
    y_te  = test["target"]

    print(f"\n  Feature count: {len(feat_cols)}")
    print(f"  X_train shape: {X_tr.shape}")
    print(f"  X_val shape:   {X_val.shape}")
    print(f"  X_test shape:  {X_te.shape}")

    print("\nTraining XGBoost...")
    model = s25_train_xgboost(X_tr, y_tr, X_val, y_val)

    proba, roc_auc, pr_auc, best_th = s25_evaluate_model(model, X_te, y_te, feat_cols)
    s25_print_risk_tiers(proba, y_te)

    print("\nstep25 complete.")


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("=" * 60)
    print("  CRE LOAN DEFAULT PREDICTION — FULL PIPELINE")
    print(f"  Started: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 60)

    panel_df    = run_step21()
    run_step22(panel_df)
    features_df = run_step24(panel_df)
    run_step25(features_df)

    print("\n" + "=" * 60)
    print("  PIPELINE COMPLETE")
    print(f"  Finished: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 60)
